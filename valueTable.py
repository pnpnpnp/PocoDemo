# -*- encoding=utf8 -*-
__author__ = "huanghonghong"

from openpyxl import load_workbook

import configparser
config = configparser.ConfigParser()
config.read("config.ini")

wb = load_workbook(config.get('valueTable','valueTablePath'),data_only=True)

ws = wb[config.get('valueTable','valueTableSheetName')]

#表坐标（因为表格位置没什么规律，暂时没什么好方法定位表格，只能手动定位）
#[第一个单元格列，第一个单元格行，表格列数，表格行数]（表结构不变的话后两个参数一般不变）
# 投放英雄
heroTableIndex = [1,20,4,4]
# 精英召唤奖励
eliteCallAwardTableIndex = [1,26,12,17]
# 礼包
superValueTableIndex = [1,48,19,6]
# 抽卡主题任务
themeTaskTableIndex = [1,69,8,6]
# 每日任务
dailyTaskTableIndex = [1,82,8,4]
# 大富翁关卡奖励
monoPolyPosTableIndex = [1,117,8,32]
# 阶段奖励
activeTaskStateAwardTableIndex = [1,155,9,5]
# 高抽特权卡
highDrawPrivilegeCardTableIndex = [1,259,3,8]
# 高抽特权卡价格
highDrawPrivilegeCardPriceIndex = [258,2]
#对应表名的字典
tableNameDict = {'投放英雄':heroTableIndex,'精英召唤奖励':eliteCallAwardTableIndex,'礼包':superValueTableIndex,
                 '抽卡主题任务':themeTaskTableIndex,'每日任务':dailyTaskTableIndex,'大富翁关卡奖励':monoPolyPosTableIndex,
                 '阶段奖励':activeTaskStateAwardTableIndex,'高抽特权卡':highDrawPrivilegeCardTableIndex}

#找需求列索引
def findIndex(colNameList,tableIndexList):
    indexList = []
    colNameListIndex = 0
    for column in range(tableIndexList[2]):
        if ws.cell(tableIndexList[1],tableIndexList[0] + column).value == colNameList[colNameListIndex]:
            indexList.append(tableIndexList[0] + column)
            colNameListIndex = colNameListIndex + 1
            if colNameListIndex == len(colNameList):
                break
    return indexList

#needColNameList必须是顺序的
#数值表里抽卡主题任务可能有同一行有相同的奖励，不合理，建议改数值表！！！
def getValueTableData(tableName,needColNameList):
    #数值表数据
    dataList = []
    #表坐标
    tableIndex = tableNameDict[tableName]
    #根据表坐标和需求列找需求列的索引
    indexList = findIndex(needColNameList,tableIndex)
    #根据需求列索引遍历表里的数据
    for i in range(tableIndex[3] - 1):
        if ws.cell(tableIndex[1] + i + 1,indexList[0]).value == None:
            break
        composeData = ''
        for j in range(len(indexList)):
            if ws.cell(tableIndex[1] + i + 1,indexList[j]).value == None:
                break
            composeData = composeData + str(ws.cell(tableIndex[1] + i + 1,indexList[j]).value) + ','
        composeData = composeData.strip(',')
        dataList.append(composeData)
    #返回数据
    return dataList

#获取指定单元格数据
def getValueTableCellData(indexList):
    cellData = ws.cell(indexList[0],indexList[1]).value
    return cellData

